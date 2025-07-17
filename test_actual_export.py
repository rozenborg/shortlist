#!/usr/bin/env python3
"""
Test the actual export endpoint that gets called when clicking Export button
"""

import requests
import json
import sys
import time

def test_export_endpoint():
    """Test the actual /api/export endpoint"""
    print("🔍 Testing Actual Export Endpoint...")
    
    base_url = "http://localhost:5001"
    
    # Test 1: Check if server is running
    try:
        response = requests.get(f"{base_url}/", timeout=5)
        print(f"✅ Server is running (status: {response.status_code})")
    except requests.exceptions.ConnectionError:
        print("❌ Server not running. Start with: python app.py")
        return False
    except Exception as e:
        print(f"❌ Server connection error: {e}")
        return False
    
    # Test 2: Check saved candidates endpoint
    try:
        response = requests.get(f"{base_url}/api/saved", timeout=10)
        if response.status_code == 200:
            saved_data = response.json()
            print(f"✅ Saved candidates endpoint works ({len(saved_data)} candidates)")
            
            if len(saved_data) == 0:
                print("⚠️  No saved candidates - nothing to export")
                return False
        else:
            print(f"❌ Saved candidates endpoint failed: {response.status_code}")
            return False
    except Exception as e:
        print(f"❌ Error getting saved candidates: {e}")
        return False
    
    # Test 3: Test the actual export endpoint
    try:
        print("🔄 Testing export endpoint...")
        response = requests.post(f"{base_url}/api/export", 
                               headers={'Content-Type': 'application/json'},
                               timeout=30)
        
        print(f"📊 Export response status: {response.status_code}")
        print(f"📊 Export response headers: {dict(response.headers)}")
        
        if response.status_code == 200:
            # Check if it's actually an Excel file
            content_type = response.headers.get('content-type', '')
            if 'spreadsheet' in content_type or 'excel' in content_type:
                file_size = len(response.content)
                print(f"✅ Export successful! Excel file created ({file_size} bytes)")
                
                # Save to test file to verify
                with open('test_export_actual.xlsx', 'wb') as f:
                    f.write(response.content)
                print("✅ Test file saved as 'test_export_actual.xlsx'")
                return True
            else:
                print(f"❌ Wrong content type: {content_type}")
                print(f"📄 Response body: {response.text[:500]}")
                return False
        elif response.status_code == 400:
            error_data = response.json()
            print(f"❌ Export failed (400): {error_data}")
            return False
        elif response.status_code == 500:
            error_data = response.json()
            print(f"❌ Server error (500): {error_data}")
            return False
        else:
            print(f"❌ Unexpected status code: {response.status_code}")
            print(f"📄 Response: {response.text}")
            return False
            
    except requests.exceptions.Timeout:
        print("❌ Export request timed out (30 seconds)")
        return False
    except Exception as e:
        print(f"❌ Export request failed: {e}")
        return False

def test_backend_export_directly():
    """Test the export function directly by calling the backend code"""
    print("\n🔍 Testing Backend Export Function Directly...")
    
    try:
        import sys
        sys.path.append('.')
        
        # Import the necessary components
        from src.candidate_service import CandidateService
        from src.manager import LLMService
        from src.resume_parser import ResumeParser
        from src.customization_service import CustomizationService
        from src.factory import get_llm_client
        from app import parse_name_from_filename
        
        # Initialize services (same as app.py)
        client = get_llm_client()
        llm_service = LLMService(client)
        resume_parser = ResumeParser()
        customization_service = CustomizationService()
        candidate_service = CandidateService(llm_service, resume_parser, customization_service)
        
        print("✅ Services initialized")
        
        # Get saved candidates (same as export function)
        saved_candidates = candidate_service.get_saved_candidates()
        print(f"✅ Got {len(saved_candidates)} saved candidates")
        
        if not saved_candidates:
            print("❌ No candidates to export")
            return False
        
        # Test name parsing with actual data
        for i, candidate in enumerate(saved_candidates[:3]):  # Test first 3
            filename = candidate.get('filename', '')
            try:
                first_name, last_name = parse_name_from_filename(filename)
                print(f"✅ Name parsing {i+1}: '{filename}' → '{first_name}', '{last_name}'")
            except Exception as e:
                print(f"❌ Name parsing failed for '{filename}': {e}")
                return False
        
        # Test Excel creation with real data
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill
            import io
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Shortlisted Candidates"
            
            # Define headers (same as app.py)
            headers = [
                'First Name', 'Last Name', 'Resume Filename', 'Nickname', 
                'Summary', 'Reservations', 'Fit Indicators', 'Achievements', 
                'Experience Distribution', 'Starred?'
            ]
            
            # Write headers
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            print("✅ Headers written")
            
            # Write candidate data (same logic as app.py)
            for row, candidate in enumerate(saved_candidates, 2):
                filename = candidate.get('filename', '')
                first_name, last_name = parse_name_from_filename(filename)
                
                # Format arrays as comma-separated strings
                reservations = ', '.join(candidate.get('reservations', []))
                fit_indicators = ', '.join(candidate.get('fit_indicators', []))
                achievements = ', '.join(candidate.get('achievements', []))
                
                # Format experience distribution
                exp_dist = candidate.get('experience_distribution', {})
                exp_text = ', '.join([f"{sector.title()}: {years}y" 
                                    for sector, years in exp_dist.items() if years > 0])
                
                # Starred status
                starred = 'TRUE' if candidate.get('is_starred', False) else ''
                
                # Write row data
                row_data = [
                    first_name,
                    last_name,
                    filename,
                    candidate.get('nickname', candidate.get('name', '')),
                    candidate.get('summary', ''),
                    reservations,
                    fit_indicators,
                    achievements,
                    exp_text,
                    starred
                ]
                
                for col, value in enumerate(row_data, 1):
                    ws.cell(row=row, column=col, value=str(value) if value else '')
            
            print(f"✅ Data written for {len(saved_candidates)} candidates")
            
            # Auto-adjust column widths (same as app.py)
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            print("✅ Column widths adjusted")
            
            # Save to BytesIO (same as app.py)
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            
            size = len(output.getvalue())
            print(f"✅ Excel file created in memory ({size} bytes)")
            
            # Test saving to actual file
            wb.save('test_backend_direct.xlsx')
            print("✅ Test file saved as 'test_backend_direct.xlsx'")
            
            return True
            
        except Exception as e:
            print(f"❌ Excel creation with real data failed: {e}")
            import traceback
            print(f"📄 Detailed error: {traceback.format_exc()}")
            return False
            
    except Exception as e:
        print(f"❌ Backend test failed: {e}")
        import traceback
        print(f"📄 Detailed error: {traceback.format_exc()}")
        return False

def main():
    print("🔍 Testing Actual Export Process")
    print("=" * 50)
    
    # Test 1: Direct backend test
    backend_success = test_backend_export_directly()
    
    print("\n" + "=" * 50)
    
    # Test 2: HTTP endpoint test
    endpoint_success = test_export_endpoint()
    
    print("\n" + "=" * 50)
    print("📊 RESULTS:")
    print(f"  Backend direct test: {'✅ PASS' if backend_success else '❌ FAIL'}")
    print(f"  HTTP endpoint test: {'✅ PASS' if endpoint_success else '❌ FAIL'}")
    
    if backend_success and endpoint_success:
        print("\n🎉 Both tests passed - export should work!")
    elif backend_success and not endpoint_success:
        print("\n🤔 Backend works but HTTP fails - likely a Flask/network issue")
        print("💡 Check: CORS settings, Flask configuration, or request headers")
    elif not backend_success:
        print("\n🚨 Backend test failed - there's an issue with the data or processing")
        print("💡 Check the detailed error above")
    
    # Cleanup test files
    import os
    for test_file in ['test_export_actual.xlsx', 'test_backend_direct.xlsx']:
        if os.path.exists(test_file):
            os.remove(test_file)
            print(f"🧹 Cleaned up {test_file}")

if __name__ == '__main__':
    main() 