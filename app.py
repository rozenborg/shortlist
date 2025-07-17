from dotenv import load_dotenv
load_dotenv()

from flask import Flask, render_template, jsonify, request, send_file
from flask_cors import CORS
import os
import json
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from src.factory import get_llm_client
from src.manager import LLMService
from src.resume_parser import ResumeParser
from src.candidate_service import CandidateService
from src.background_processor import BackgroundProcessor
from src.customization_service import CustomizationService

app = Flask(__name__)
CORS(app)

# Initialize services with real-time processing and smart retry logic
client = get_llm_client()
llm_service = LLMService(client)
resume_parser = ResumeParser()
customization_service = CustomizationService()
candidate_service = CandidateService(llm_service, resume_parser, customization_service)
background_processor = BackgroundProcessor(candidate_service, resume_parser, llm_service)

print("🚀 Resume processing with real-time updates and smart retry logic enabled")

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/api/candidates', methods=['GET'])
def get_candidates():
    """Get all candidates from the candidates folder"""
    candidates = candidate_service.get_all_candidates()
    return jsonify(candidates)

@app.route('/api/candidate/<candidate_id>', methods=['GET'])
def get_candidate(candidate_id):
    """Get a specific candidate by ID"""
    candidate = candidate_service.get_candidate(candidate_id)
    if candidate:
        return jsonify(candidate)
    return jsonify({'error': 'Candidate not found'}), 404

@app.route('/api/swipe', methods=['POST'])
def handle_swipe():
    """Handle swipe decision (left/right)"""
    data = request.json
    candidate_id = data.get('candidate_id')
    decision = data.get('decision')  # 'pass' or 'save'
    
    result = candidate_service.save_decision(candidate_id, decision)
    return jsonify(result)

@app.route('/api/saved', methods=['GET'])
def get_saved_candidates():
    """Get all saved candidates"""
    saved = candidate_service.get_saved_candidates()
    return jsonify(saved)

@app.route('/api/saved/reorder', methods=['POST'])
def reorder_saved_candidates():
    """Update the custom order of saved candidates"""
    data = request.json
    ordered_ids = data.get('ordered_ids', [])
    
    if not ordered_ids:
        return jsonify({'error': 'No ordered_ids provided'}), 400
    
    result = candidate_service.update_candidate_order(ordered_ids)
    return jsonify(result)

@app.route('/api/process/start', methods=['POST'])
def start_background_processing():
    """Start background processing of all resumes"""
    background_processor.start_background_processing()
    return jsonify({'status': 'started'})

@app.route('/api/process/status', methods=['GET'])
def get_processing_status():
    """Get current processing status"""
    status = background_processor.get_status()
    return jsonify(status)

@app.route('/api/process/stats', methods=['GET'])
def get_processing_stats():
    """Get detailed processing statistics"""
    stats = candidate_service.get_processing_stats()
    return jsonify(stats)

@app.route('/api/candidates/ready', methods=['GET'])
def get_ready_candidates():
    """Get only candidates that are ready for review"""
    candidates = candidate_service.get_ready_candidates()
    return jsonify(candidates)

@app.route('/api/candidates/processing', methods=['GET'])
def get_processing_candidates():
    """Get candidates currently being processed"""
    candidates = candidate_service.get_processing_candidates()
    return jsonify(candidates)

@app.route('/api/candidates/newly-processed', methods=['GET'])
def get_newly_processed_candidates():
    """Get candidates that were recently processed (for real-time updates)"""
    candidates = candidate_service.get_newly_processed_candidates(background_processor)
    return jsonify(candidates)

@app.route('/api/process/failed', methods=['GET'])
def get_failed_candidates():
    """Get candidates that failed after max retries"""
    failed = background_processor.get_failed_candidates()
    return jsonify(failed)

@app.route('/api/process/format-issues', methods=['GET'])
def get_format_issues():
    """Get candidates with formatting issues in retry queues"""
    status = background_processor.get_status()
    format_queue_size = status['retry_queues'].get('format_retry', 0)
    
    # Get detailed information about formatting issues
    format_issues = []
    for candidate in background_processor.retry_queues.get('format_retry', []):
        last_response = candidate.get('_last_response', {})
        quality_info = last_response.get('_quality_info', {})
        
        format_issues.append({
            'id': candidate['id'],
            'filename': candidate['filename'],
            'name': candidate.get('name', 'Unknown'),
            'retry_count': background_processor.retry_counts.get(candidate['id'], 0),
            'quality_issues': quality_info.get('details', []),
            'quality_score': quality_info.get('quality_score', 0),
            'last_error': quality_info.get('reason', 'Unknown formatting issue')
        })
    
    return jsonify({
        'format_queue_size': format_queue_size,
        'format_issues': format_issues
    })

@app.route('/api/export-failed', methods=['POST'])
def export_failed_candidates():
    """Export failed candidates to Excel file"""
    try:
        # Get failed candidates
        failed_candidates = background_processor.get_failed_candidates()
        
        if not failed_candidates:
            return jsonify({'error': 'No failed candidates to export'}), 400
        
        # Create workbook and worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Failed to Process"
        
        # Define headers
        headers = [
            'Resume Filename', 'Candidate Name', 'Error Type', 
            'Error Message', 'Retry Count', 'Failed At'
        ]
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Write failed candidate data
        for row, candidate in enumerate(failed_candidates, 2):
            # Get retry count
            retry_count = background_processor.retry_counts.get(candidate['id'], 0)
            
            # Get timestamp if available
            failed_at = candidate.get('failed_at', 'Unknown')
            
            # Write row data
            row_data = [
                candidate.get('filename', 'Unknown'),
                candidate.get('name', 'Unknown'),
                candidate.get('error_type', 'Unknown'),
                candidate.get('error', 'Unknown error'),
                retry_count,
                failed_at
            ]
            
            for col, value in enumerate(row_data, 1):
                ws.cell(row=row, column=col, value=value)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 60)  # Cap at 60 characters for error messages
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return send_file(
            output,
            as_attachment=True,
            download_name='failed_candidates.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        print(f"Error exporting failed candidates: {e}")
        return jsonify({'error': 'Failed to export failed candidates'}), 500

@app.route('/api/process/retry/<candidate_id>', methods=['POST'])
def retry_failed_candidate(candidate_id):
    """Manually retry a failed candidate"""
    success = background_processor.retry_failed_candidate(candidate_id)
    
    if success:
        return jsonify({'success': True, 'message': f'Candidate {candidate_id} moved to retry queue'})
    else:
        return jsonify({'success': False, 'message': 'Candidate not found in failed queue'}), 404

@app.route('/api/process/config', methods=['GET'])
def get_processing_config():
    """Get current processing configuration"""
    status = background_processor.get_status()
    return jsonify(status.get('config', {}))

@app.route('/api/process/config', methods=['POST'])
def update_processing_config():
    """Update processing configuration"""
    data = request.json
    
    # Update configuration in background processor
    config_updates = {}
    
    if 'quick_timeout' in data:
        config_updates['quick_timeout'] = int(data['quick_timeout'])
    if 'long_timeout' in data:
        config_updates['long_timeout'] = int(data['long_timeout'])
    if 'max_retries' in data:
        config_updates['max_retries'] = int(data['max_retries'])
    if 'batch_size' in data:
        config_updates['batch_size'] = int(data['batch_size'])
    
    # Apply configuration updates
    background_processor.config.update(config_updates)
    
    return jsonify({'success': True, 'updated_config': config_updates})

@app.route('/api/process/batch', methods=['POST'])
def process_batch():
    """Process a specific batch of candidates immediately"""
    data = request.json
    candidate_ids = data.get('candidate_ids', [])
    
    if not candidate_ids:
        return jsonify({'error': 'No candidate IDs provided'}), 400
    
    results = background_processor.force_process_batch(candidate_ids)
    return jsonify({
        'processed_count': len(results),
        'results': results
    })

@app.route('/api/job-description', methods=['GET', 'POST'])
def job_description():
    """Handle job description for candidate screening"""
    if request.method == 'POST':
        data = request.json
        job_description = data.get('job_description', '')
        
        if not job_description.strip():
            return jsonify({'error': 'Job description is required'}), 400
            
        result = customization_service.update_settings(job_description)
        
        # Clear cache to force re-analysis with new job description
        if os.path.exists(candidate_service.summaries_cache):
            os.remove(candidate_service.summaries_cache)
            
        # Clear in-memory summaries cache
        candidate_service.summaries = {}
        candidate_service._load_data()  # Reload empty data from disk
            
        # Start processing resumes with the new job description
        background_processor.start_background_processing()
        
        return jsonify(result)
    else:
        # GET request - check if job description exists
        settings = customization_service.get_settings()
        has_job_description = bool(settings.get('job_description', '').strip())
        return jsonify({
            'has_job_description': has_job_description,
            'job_description': settings.get('job_description', '')
        })

@app.route('/api/undo', methods=['POST'])
def undo_swipe():
    """Undo the last swipe"""
    result = candidate_service.undo_last_swipe()
    return jsonify(result)

@app.route('/api/restart', methods=['POST'])
def restart_session():
    """Restart the session"""
    result = candidate_service.restart_session()
    # Don't clear the job description, just the decisions
    background_processor.start_background_processing()
    return jsonify(result)

@app.route('/api/passed', methods=['GET'])
def get_passed_candidates():
    """Get all passed candidates"""
    passed_candidates = candidate_service.get_passed_candidates()
    return jsonify(passed_candidates)

@app.route('/api/modify-decision', methods=['POST'])
def modify_decision():
    """Modify an existing decision for a candidate"""
    data = request.json
    candidate_id = data.get('candidate_id')
    new_decision = data.get('new_decision')  # 'save', 'pass', 'star', or 'unreviewed'
    
    if not candidate_id or not new_decision:
        return jsonify({'success': False, 'message': 'Missing candidate_id or new_decision'}), 400
    
    valid_decisions = ['save', 'pass', 'star', 'unreviewed']
    if new_decision not in valid_decisions:
        return jsonify({'success': False, 'message': f'Invalid decision. Must be one of: {valid_decisions}'}), 400
    
    result = candidate_service.modify_decision(candidate_id, new_decision)
    return jsonify(result)

@app.route('/api/export', methods=['POST'])
def export_candidates():
    """Export saved candidates to Excel file"""
    try:
        # Get saved candidates
        saved_candidates = candidate_service.get_saved_candidates()
        
        if not saved_candidates:
            return jsonify({'error': 'No candidates to export'}), 400
        
        # Create workbook and worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Shortlisted Candidates"
        
        # Define headers
        headers = [
            'First Name', 'Last Name', 'Resume Filename', 'Nickname', 
            'Summary', 'Reservations', 'Fit Indicators', 'Achievements', 
            'Experience Distribution', 'Starred?'
        ]
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Write candidate data
        for row, candidate in enumerate(saved_candidates, 2):
            # Parse name from filename
            filename = candidate.get('filename', '')
            first_name, last_name = parse_name_from_filename(filename)
            
            # Format arrays as comma-separated strings
            reservations = ', '.join(candidate.get('reservations', []))
            fit_indicators = ', '.join(candidate.get('fit_indicators', []))
            achievements = ', '.join(candidate.get('achievements', []))
            
            # Format experience distribution
            exp_dist = candidate.get('experience_distribution', {})
            exp_text = ', '.join([f"{sector.title()}: {years}y" 
                                for sector, years in exp_dist.items() 
                                if (isinstance(years, (int, float)) and years > 0) or 
                                   (isinstance(years, str) and years.isdigit() and int(years) > 0)])
            
            # Starred status
            starred = 'TRUE' if candidate.get('is_starred', False) else ''
            
            # Write row data
            row_data = [
                first_name,
                last_name,
                filename,
                candidate.get('nickname', candidate.get('name', '')),
                candidate.get('summary', ''),
                reservations,
                fit_indicators,
                achievements,
                exp_text,
                starred
            ]
            
            for col, value in enumerate(row_data, 1):
                ws.cell(row=row, column=col, value=value)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return send_file(
            output,
            as_attachment=True,
            download_name='shortlisted_candidates.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        print(f"Error exporting candidates: {e}")
        return jsonify({'error': 'Failed to export candidates'}), 500

def parse_name_from_filename(filename):
    """Parse first and last name from resume filename"""
    if not filename:
        return 'Unknown', 'Name'
    
    # Remove file extension
    name_part = os.path.splitext(filename)[0]
    
    import re
    # Pattern to match: "FirstName LastName(s) ID RESUME"
    # This regex looks for the pattern where after the name, there's a space followed by 
    # an ID (mix of letters/numbers) and then "RESUME"
    # It preserves hyphenated last names like "Smith-Jones"
    match = re.match(r'^(.+?)\s+[a-zA-Z0-9_-]+\s+RESUME.*$', name_part, re.IGNORECASE)
    
    if match:
        # Extract the name part before the ID
        name_part = match.group(1).strip()
    else:
        # Fallback: just remove " RESUME" and anything after it
        name_part = re.sub(r'\s+RESUME.*$', '', name_part, flags=re.IGNORECASE)
    
    # Split by spaces to get first and last name
    parts = name_part.strip().split()
    
    if len(parts) >= 2:
        first_name = parts[0]
        # Join all remaining parts as last name (handles middle names and hyphenated names)
        last_name = ' '.join(parts[1:])
    elif len(parts) == 1:
        first_name = parts[0]
        last_name = ''
    else:
        first_name = 'Unknown'
        last_name = 'Name'
    
    return first_name, last_name

if __name__ == '__main__':
    # Create necessary directories
    os.makedirs('candidates', exist_ok=True)
    os.makedirs('data', exist_ok=True)
    os.makedirs('templates', exist_ok=True)
    os.makedirs('static', exist_ok=True)
    os.makedirs('static/css', exist_ok=True)
    os.makedirs('static/js', exist_ok=True)
    
    # Start background processing automatically
    print("Starting background processing...")
    background_processor.start_background_processing()
    
    app.run(host='127.0.0.1', port=5001, debug=False) 