#!/usr/bin/env python3
"""
Debug version of the export function to reveal the real error
This bypasses the generic error handling to show exactly what's failing
"""

import sys
import traceback

def debug_export_function():
    """Run the export function with detailed error reporting"""
    print("🔍 Debug Export Function - Revealing Real Errors")
    print("=" * 60)
    
    try:
        # Import all necessary components (same as app.py)
        sys.path.append('.')
        
        from src.candidate_service import CandidateService
        from src.manager import LLMService
        from src.resume_parser import ResumeParser
        from src.customization_service import CustomizationService
        from src.factory import get_llm_client
        from app import parse_name_from_filename
        
        print("✅ Imports successful")
        
        # Initialize services (same as app.py)
        client = get_llm_client()
        llm_service = LLMService(client)
        resume_parser = ResumeParser()
        customization_service = CustomizationService()
        candidate_service = CandidateService(llm_service, resume_parser, customization_service)
        
        print("✅ Services initialized")
        
        # Step 1: Get saved candidates
        print("\n🔄 Step 1: Getting saved candidates...")
        saved_candidates = candidate_service.get_saved_candidates()
        print(f"✅ Got {len(saved_candidates)} saved candidates")
        
        if not saved_candidates:
            print("❌ No candidates to export - this would return 400 error")
            return
        
        # Step 2: Import Excel libraries
        print("\n🔄 Step 2: Importing Excel libraries...")
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        import io
        print("✅ Excel libraries imported")
        
        # Step 3: Create workbook and worksheet
        print("\n🔄 Step 3: Creating workbook...")
        wb = Workbook()
        ws = wb.active
        ws.title = "Shortlisted Candidates"
        print("✅ Workbook created")
        
        # Step 4: Define and write headers
        print("\n🔄 Step 4: Writing headers...")
        headers = [
            'First Name', 'Last Name', 'Resume Filename', 'Nickname', 
            'Summary', 'Reservations', 'Fit Indicators', 'Achievements', 
            'Experience Distribution', 'Starred?'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        print(f"✅ Headers written ({len(headers)} columns)")
        
        # Step 5: Process each candidate
        print(f"\n🔄 Step 5: Processing {len(saved_candidates)} candidates...")
        
        for row_idx, candidate in enumerate(saved_candidates, 2):
            print(f"\n  Processing candidate {row_idx-1}/{len(saved_candidates)}:")
            
            # Parse name from filename
            filename = candidate.get('filename', '')
            print(f"    Filename: '{filename}'")
            
            try:
                first_name, last_name = parse_name_from_filename(filename)
                print(f"    Name parsing: '{first_name}', '{last_name}'")
            except Exception as e:
                print(f"    ❌ Name parsing failed: {e}")
                raise
            
            # Format arrays as comma-separated strings
            try:
                reservations = candidate.get('reservations', [])
                print(f"    Reservations type: {type(reservations)}, value: {reservations}")
                reservations_str = ', '.join(reservations) if reservations else ''
                
                fit_indicators = candidate.get('fit_indicators', [])
                print(f"    Fit indicators type: {type(fit_indicators)}")
                fit_indicators_str = ', '.join(fit_indicators) if fit_indicators else ''
                
                achievements = candidate.get('achievements', [])
                print(f"    Achievements type: {type(achievements)}")
                achievements_str = ', '.join(achievements) if achievements else ''
                
                print("    ✅ Array formatting successful")
                
            except Exception as e:
                print(f"    ❌ Array formatting failed: {e}")
                print(f"    Candidate data: {candidate}")
                raise
            
            # Format experience distribution
            try:
                exp_dist = candidate.get('experience_distribution', {})
                print(f"    Experience dist type: {type(exp_dist)}, value: {exp_dist}")
                
                if isinstance(exp_dist, dict):
                    exp_text = ', '.join([f"{sector.title()}: {years}y" 
                                        for sector, years in exp_dist.items() 
                                        if (isinstance(years, (int, float)) and years > 0) or 
                                           (isinstance(years, str) and years.isdigit() and int(years) > 0)])
                else:
                    print(f"    ⚠️ Experience distribution is not a dict: {exp_dist}")
                    exp_text = str(exp_dist) if exp_dist else ''
                
                print(f"    Experience text: '{exp_text}'")
                
            except Exception as e:
                print(f"    ❌ Experience distribution formatting failed: {e}")
                raise
            
            # Starred status
            try:
                starred = 'TRUE' if candidate.get('is_starred', False) else ''
                print(f"    Starred: {starred}")
            except Exception as e:
                print(f"    ❌ Starred status failed: {e}")
                raise
            
            # Create row data
            try:
                row_data = [
                    first_name,
                    last_name,
                    filename,
                    candidate.get('nickname', candidate.get('name', '')),
                    candidate.get('summary', ''),
                    reservations_str,
                    fit_indicators_str,
                    achievements_str,
                    exp_text,
                    starred
                ]
                
                print(f"    Row data length: {len(row_data)}")
                
                # Write row data to Excel
                for col, value in enumerate(row_data, 1):
                    cell_value = str(value) if value is not None else ''
                    ws.cell(row=row_idx, column=col, value=cell_value)
                
                print(f"    ✅ Candidate {row_idx-1} written to Excel")
                
            except Exception as e:
                print(f"    ❌ Writing row data failed: {e}")
                print(f"    Row data: {row_data}")
                raise
        
        print(f"\n✅ All {len(saved_candidates)} candidates processed")
        
        # Step 6: Auto-adjust column widths
        print("\n🔄 Step 6: Adjusting column widths...")
        try:
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            print("✅ Column widths adjusted")
            
        except Exception as e:
            print(f"❌ Column width adjustment failed: {e}")
            raise
        
        # Step 7: Save to BytesIO
        print("\n🔄 Step 7: Saving to BytesIO...")
        try:
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            
            size = len(output.getvalue())
            print(f"✅ Excel file created successfully ({size} bytes)")
            
            # Test saving to actual file
            wb.save('debug_export_test.xlsx')
            print("✅ Debug file saved as 'debug_export_test.xlsx'")
            
            return True
            
        except Exception as e:
            print(f"❌ Saving to BytesIO failed: {e}")
            raise
            
    except Exception as e:
        print(f"\n💥 REAL ERROR FOUND: {e}")
        print(f"📄 Error type: {type(e).__name__}")
        print(f"📄 Full traceback:")
        traceback.print_exc()
        return False

if __name__ == '__main__':
    print("🐛 Debug Export Function")
    print("This will show the exact error that's being hidden by the generic message")
    print("=" * 80)
    
    success = debug_export_function()
    
    print("\n" + "=" * 80)
    if success:
        print("🎉 Debug export completed successfully!")
        print("💡 The issue might be Flask/HTTP related, not the export logic")
    else:
        print("🚨 Debug export failed - the real error is shown above")
        print("💡 Fix this error to resolve the export issue") 