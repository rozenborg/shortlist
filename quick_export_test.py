#!/usr/bin/env python3
"""
Quick test to check if Excel export will work
"""

def test_export_requirements():
    print("🔍 Testing Export Requirements...")
    
    # Test 1: openpyxl availability
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        import io
        print("✅ openpyxl imported successfully")
    except ImportError as e:
        print(f"❌ openpyxl import failed: {e}")
        print("💡 Solution: pip install openpyxl")
        return False
    
    # Test 2: Basic Excel creation
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Test"
        
        # Add test data with formatting
        cell = ws.cell(row=1, column=1, value="Test")
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
        
        # Save to memory
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        size = len(output.getvalue())
        print(f"✅ Excel file created successfully ({size} bytes)")
        return True
        
    except Exception as e:
        print(f"❌ Excel creation failed: {e}")
        return False

if __name__ == '__main__':
    if test_export_requirements():
        print("🎉 Export should work on this system!")
    else:
        print("🚨 Export will fail - fix the issues above") 