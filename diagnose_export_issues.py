#!/usr/bin/env python3
"""
Export Diagnostics for Swipe Application
This script helps diagnose why Excel export might be failing.
"""

import os
import sys
import json
from datetime import datetime

def print_header(title):
    print(f"\n{'='*60}")
    print(f"🔍 {title}")
    print('='*60)

def print_success(message):
    print(f"✅ {message}")

def print_warning(message):
    print(f"⚠️  {message}")

def print_error(message):
    print(f"❌ {message}")

def print_info(message):
    print(f"ℹ️  {message}")

def test_imports():
    """Test if all required libraries for export are available"""
    print_header("Testing Required Import Libraries")
    
    required_libraries = [
        ('openpyxl', 'openpyxl'),
        ('openpyxl.Workbook', 'openpyxl'),
        ('openpyxl.styles.Font', 'openpyxl.styles'),
        ('openpyxl.styles.Alignment', 'openpyxl.styles'),  
        ('openpyxl.styles.PatternFill', 'openpyxl.styles'),
        ('io', 'io'),
        ('flask', 'flask'),
    ]
    
    results = {}
    
    for lib_name, import_path in required_libraries:
        try:
            if lib_name == 'openpyxl.Workbook':
                from openpyxl import Workbook
                print_success(f"{lib_name}: Available")
                results[lib_name] = True
            elif lib_name == 'openpyxl.styles.Font':
                from openpyxl.styles import Font
                print_success(f"{lib_name}: Available")
                results[lib_name] = True
            elif lib_name == 'openpyxl.styles.Alignment':
                from openpyxl.styles import Alignment
                print_success(f"{lib_name}: Available") 
                results[lib_name] = True
            elif lib_name == 'openpyxl.styles.PatternFill':
                from openpyxl.styles import PatternFill
                print_success(f"{lib_name}: Available")
                results[lib_name] = True
            elif lib_name == 'io':
                import io
                print_success(f"{lib_name}: Available")
                results[lib_name] = True
            elif lib_name == 'flask':
                import flask
                print_success(f"Flask: {flask.__version__}")
                results[lib_name] = True
            else:
                __import__(import_path)
                print_success(f"{lib_name}: Available")
                results[lib_name] = True
                
        except ImportError as e:
            print_error(f"{lib_name}: Missing - {e}")
            results[lib_name] = False
        except Exception as e:
            print_error(f"{lib_name}: Error - {e}")
            results[lib_name] = False
    
    return results

def test_data_availability():
    """Test if there are candidates to export"""
    print_header("Testing Data Availability")
    
    data_folder = 'data'
    decisions_file = os.path.join(data_folder, 'decisions.json')
    
    if not os.path.exists(data_folder):
        print_error(f"Data folder '{data_folder}' not found")
        return False
    
    if not os.path.exists(decisions_file):
        print_warning(f"Decisions file '{decisions_file}' not found - no candidates saved yet")
        return False
    
    try:
        with open(decisions_file, 'r') as f:
            decisions = json.load(f)
        
        # Updated to handle the proper format: decisions = {'saved': [...], 'starred': [...], ...}
        saved_list = decisions.get('saved', [])
        starred_list = decisions.get('starred', [])
        passed_list = decisions.get('passed', [])
        
        # Count IDs from saved and starred lists (they may overlap)
        saved_ids = [item['id'] for item in saved_list]
        starred_ids = [item['id'] for item in starred_list]
        all_saved_or_starred_ids = list(set(saved_ids + starred_ids))  # Remove duplicates
        
        print_info(f"Total saved: {len(saved_list)}")
        print_info(f"Total starred: {len(starred_list)}")
        print_info(f"Total passed: {len(passed_list)}")
        print_info(f"Unique saved/starred candidates: {len(all_saved_or_starred_ids)}")
        
        if len(all_saved_or_starred_ids) == 0:
            print_warning("No saved or starred candidates found - nothing to export")
            return False
        else:
            print_success(f"Found {len(all_saved_or_starred_ids)} candidates to export")
            return True
            
    except Exception as e:
        print_error(f"Error reading decisions file: {e}")
        import traceback
        print_error(f"Detailed error: {traceback.format_exc()}")
        return False

def test_excel_creation():
    """Test if we can create a simple Excel file"""
    print_header("Testing Excel File Creation")
    
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        import io
        
        # Create a simple workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Test Export"
        
        # Add some test data
        headers = ['Name', 'Test Data']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Add test row
        ws.cell(row=2, column=1, value="Test Candidate")
        ws.cell(row=2, column=2, value="Sample data")
        
        # Try to save to BytesIO (same as export function)
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        size = len(output.getvalue())
        print_success(f"Excel file created successfully ({size} bytes)")
        
        # Try to save to actual file for testing
        test_filename = 'test_export.xlsx'
        try:
            wb.save(test_filename)
            print_success(f"Test file '{test_filename}' saved successfully")
            
            # Clean up
            if os.path.exists(test_filename):
                os.remove(test_filename)
                print_info("Test file cleaned up")
                
        except Exception as e:
            print_error(f"Cannot save to file system: {e}")
            return False
            
        return True
        
    except Exception as e:
        print_error(f"Excel creation failed: {e}")
        return False

def test_parse_name_function():
    """Test the parse_name_from_filename function"""
    print_header("Testing Name Parsing Function")
    
    try:
        # Import the function from app.py
        sys.path.append('.')
        from app import parse_name_from_filename
        
        test_filenames = [
            'John Doe 12345 RESUME.pdf',
            'Jane Smith_67890_RESUME.docx', 
            'Alex Johnson RESUME.txt',
            'Michael Rosenberg_300071945857786_2025-07-02T22:05:44+00:00_RESUME.pdf',
            'test_file.pdf'
        ]
        
        for filename in test_filenames:
            first, last = parse_name_from_filename(filename)
            print_success(f"'{filename}' → First: '{first}', Last: '{last}'")
        
        return True
        
    except Exception as e:
        print_error(f"Name parsing test failed: {e}")
        return False

def test_candidate_service():
    """Test if we can load the candidate service and get saved candidates"""
    print_header("Testing Candidate Service")
    
    try:
        sys.path.append('.')
        from src.candidate_service import CandidateService
        from src.manager import LLMService
        from src.resume_parser import ResumeParser
        from src.customization_service import CustomizationService
        from src.factory import get_llm_client
        
        # Initialize services (same as app.py)
        client = get_llm_client()
        llm_service = LLMService(client)
        resume_parser = ResumeParser()
        customization_service = CustomizationService()
        candidate_service = CandidateService(llm_service, resume_parser, customization_service)
        
        # Try to get saved candidates
        saved_candidates = candidate_service.get_saved_candidates()
        print_success(f"Candidate service loaded successfully")
        print_info(f"Found {len(saved_candidates)} saved candidates")
        
        if len(saved_candidates) > 0:
            # Show sample data structure
            sample = saved_candidates[0]
            print_info(f"Sample candidate keys: {list(sample.keys())}")
        
        return True, saved_candidates
        
    except Exception as e:
        print_error(f"Candidate service test failed: {e}")
        return False, []

def test_full_export_process():
    """Test the complete export process"""
    print_header("Testing Complete Export Process")
    
    try:
        # Test candidate service first
        success, saved_candidates = test_candidate_service()
        if not success or len(saved_candidates) == 0:
            print_warning("Cannot test full export - no candidates available")
            return False
        
        # Import everything needed for export
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        import io
        from app import parse_name_from_filename
        
        # Simulate the export process (same logic as app.py)
        wb = Workbook()
        ws = wb.active
        ws.title = "Shortlisted Candidates"
        
        # Define headers (same as app.py)
        headers = [
            'First Name', 'Last Name', 'Resume Filename', 'Nickname', 
            'Summary', 'Reservations', 'Fit Indicators', 'Achievements', 
            'Experience Distribution', 'Starred?'
        ]
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Write candidate data (same logic as app.py)
        for row, candidate in enumerate(saved_candidates[:3], 2):  # Test with first 3 candidates
            filename = candidate.get('filename', '')
            first_name, last_name = parse_name_from_filename(filename)
            
            # Format arrays as comma-separated strings
            reservations = ', '.join(candidate.get('reservations', []))
            fit_indicators = ', '.join(candidate.get('fit_indicators', []))
            achievements = ', '.join(candidate.get('achievements', []))
            
            # Format experience distribution
            exp_dist = candidate.get('experience_distribution', {})
            exp_text = ', '.join([f"{sector.title()}: {years}y" 
                                for sector, years in exp_dist.items() if years > 0])
            
            # Starred status
            starred = 'TRUE' if candidate.get('is_starred', False) else ''
            
            # Write row data
            row_data = [
                first_name,
                last_name,
                filename,
                candidate.get('nickname', candidate.get('name', '')),
                candidate.get('summary', ''),
                reservations,
                fit_indicators,
                achievements,
                exp_text,
                starred
            ]
            
            for col, value in enumerate(row_data, 1):
                ws.cell(row=row, column=col, value=value)
        
        # Auto-adjust column widths (same as app.py)
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to BytesIO (same as app.py)
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        size = len(output.getvalue())
        print_success(f"Full export process completed successfully!")
        print_info(f"Generated Excel file: {size} bytes")
        print_info(f"Processed {len(saved_candidates[:3])} candidates")
        
        return True
        
    except Exception as e:
        print_error(f"Full export process failed: {e}")
        import traceback
        print_error(f"Detailed error: {traceback.format_exc()}")
        return False

def main():
    """Run all diagnostic tests"""
    print_header("Export Diagnostics for Swipe Application")
    print(f"🕒 Started at: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"🐍 Python version: {sys.version}")
    print(f"📁 Working directory: {os.getcwd()}")
    
    test_results = {}
    
    # Test 1: Required imports
    test_results['imports'] = test_imports()
    
    # Test 2: Data availability
    test_results['data'] = test_data_availability()
    
    # Test 3: Basic Excel creation
    test_results['excel_creation'] = test_excel_creation()
    
    # Test 4: Name parsing function
    test_results['name_parsing'] = test_parse_name_function()
    
    # Test 5: Full export process (if other tests pass)
    if all([test_results['imports'], test_results['data'], test_results['excel_creation']]):
        test_results['full_export'] = test_full_export_process()
    else:
        print_warning("Skipping full export test due to previous failures")
        test_results['full_export'] = False
    
    # Summary
    print_header("DIAGNOSTIC SUMMARY")
    
    passed_tests = sum(1 for result in test_results.values() if result)
    total_tests = len(test_results)
    
    print(f"📊 Tests passed: {passed_tests}/{total_tests}")
    
    for test_name, result in test_results.items():
        status = "✅ PASS" if result else "❌ FAIL"
        print(f"  {test_name}: {status}")
    
    if test_results['full_export']:
        print_success("\n🎉 Export functionality should work correctly!")
    else:
        print_error("\n🚨 Export functionality has issues. Check the failures above.")
        
        # Provide specific guidance
        if not test_results['imports']:
            print_error("👉 SOLUTION: Install missing libraries with: pip install openpyxl")
        if not test_results['data']:
            print_warning("👉 NOTE: Save some candidates first before testing export")
        if not test_results['excel_creation']:
            print_error("👉 ISSUE: Excel creation failed - check permissions and disk space")

if __name__ == '__main__':
    main() 